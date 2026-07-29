"""
Generic site-wide "Import from Excel (.xlsx)" admin feature.

Pairs with ``admin_export.py``: the user exports any model to .xlsx, edits
the file in Excel, then re-imports it. Rows are matched by the ``Id`` column:

  * If ``Id`` matches an existing PK   → UPDATE that row
  * If ``Id`` is empty / not found     → CREATE a new row

Importing this module monkey-patches ``admin.ModelAdmin.get_urls`` to add a
per-model URL (``import-excel/``) on every registered admin, and works with
the project-level ``templates/admin/change_list.html`` override that adds the
"Import from Excel" button next to "Add" on every changelist.

Field handling on import (mirrors the export rules in reverse):

  * Empty cell on UPDATE → no change to that field
  * Empty cell on CREATE → field default is used
  * Choices field        → accepts the human-readable display label OR the raw value
  * ForeignKey / O2O     → accepts the FK's PK (numeric) OR str(related_obj)
  * ManyToManyField      → comma-separated list of PKs or str() values
  * JSONField            → JSON string (parsed via json.loads)
  * DateTimeField        → ``YYYY-MM-DD HH:MM:SS`` (also tolerates ``YYYY-MM-DDTHH:MM:SS``)
  * DateField            → ``YYYY-MM-DD``
  * TimeField            → ``HH:MM:SS``
  * DecimalField         → string parsed via ``Decimal()``
  * BooleanField         → ``true/false/1/0/yes/no/y/n/t/f`` (case-insensitive)
  * FileField/ImageField → SKIPPED (URLs in the export aren't re-importable as files)
  * ``password``         → SKIPPED (sensitive)
  * ``created_at`` / ``updated_at`` → SKIPPED (auto-managed)
"""
import json
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.db.models import (
    BooleanField,
    DateField,
    DateTimeField,
    DecimalField,
    FileField,
    FloatField,
    ForeignKey,
    ImageField,
    IntegerField,
    JSONField,
    ManyToManyField,
    OneToOneField,
    TimeField,
)
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse


# Fields that should always be skipped on import (auto-managed or sensitive)
_SKIP_FIELDS = {"password", "created_at", "updated_at"}


# ── Cell parsing ─────────────────────────────────────────────────────────────

def _parse_choice(field, str_val):
    """Reverse-lookup a choices field: accept display label OR raw value."""
    for raw_val, display in field.choices:
        if str_val == str(display) or str_val == str(raw_val):
            return raw_val
    raise ValueError(
        f"'{str_val}' is not a valid choice for '{field.name}'. "
        f"Allowed: {[str(d) for _, d in field.choices]}"
    )


def _resolve_related(related_model, str_val, field_name):
    """Find a related instance by PK (preferred) or by str() match."""
    # Try numeric PK first — fastest, unambiguous
    try:
        pk = int(str(str_val).strip())
    except (ValueError, TypeError):
        pk = None
    if pk is not None:
        try:
            return related_model._default_manager.get(pk=pk)
        except related_model.DoesNotExist:
            pass

    # Fall back to str() match (slower; ambiguous strs return the first match)
    needle = str(str_val).strip()
    for obj in related_model._default_manager.all().iterator():
        if str(obj) == needle:
            return obj
    raise ValueError(
        f"Related object '{str_val}' not found for field '{field_name}'."
    )


def _parse_value(field, value):
    """Convert a single Excel cell back to a Python value for a model field."""
    # openpyxl returns None for blank cells
    if value is None:
        return None
    str_val = str(value).strip()
    if str_val == "":
        return None

    # Choices — check FIRST, before type-specific parsing, because choice
    # fields are still IntegerField / CharField underneath.
    if getattr(field, "choices", None):
        return _parse_choice(field, str_val)

    if isinstance(field, (ForeignKey, OneToOneField)):
        return _resolve_related(field.related_model, str_val, field.name)

    if isinstance(field, JSONField):
        try:
            return json.loads(str_val)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON for '{field.name}': {e}") from e

    if isinstance(field, DateTimeField):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(str_val, fmt)
            except ValueError:
                continue
        # openpyxl may return a datetime object directly
        if isinstance(value, datetime):
            return value
        raise ValueError(f"Invalid datetime for '{field.name}': '{str_val}'")

    if isinstance(field, DateField):
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        try:
            return datetime.strptime(str_val, "%Y-%m-%d").date()
        except ValueError as e:
            raise ValueError(f"Invalid date for '{field.name}': '{str_val}'") from e

    if isinstance(field, TimeField):
        if isinstance(value, time):
            return value
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(str_val, fmt).time()
            except ValueError:
                continue
        raise ValueError(f"Invalid time for '{field.name}': '{str_val}'")

    if isinstance(field, DecimalField):
        try:
            return Decimal(str_val)
        except InvalidOperation as e:
            raise ValueError(f"Invalid decimal for '{field.name}': '{str_val}'") from e

    if isinstance(field, BooleanField):
        return str_val.lower() in ("true", "1", "yes", "y", "t")

    if isinstance(field, IntegerField):
        try:
            return int(str_val)
        except ValueError as e:
            raise ValueError(f"Invalid integer for '{field.name}': '{str_val}'") from e

    if isinstance(field, FloatField):
        try:
            return float(str_val)
        except ValueError as e:
            raise ValueError(f"Invalid number for '{field.name}': '{str_val}'") from e

    return str_val


def _parse_m2m(field, value):
    """Parse a M2M cell (comma-separated) into a list of related objects."""
    if value is None or str(value).strip() == "":
        return []
    items = [s.strip() for s in str(value).split(",") if s.strip()]
    return [
        _resolve_related(field.related_model, item, field.name) for item in items
    ]


# ── Importer ─────────────────────────────────────────────────────────────────

class ImportError(Exception):
    """Raised for fatal import failures (e.g. unreadable file, missing headers)."""


def _build_field_map(model):
    """
    Return ``{header_string: field}`` for the model.

    Accepts both ``verbose_name.title()`` (the format produced by the export)
    AND the raw field name (so users can simplify column headers if they want).
    """
    opts = model._meta
    by_header = {}
    for f in list(opts.fields) + list(opts.many_to_many):
        by_header[str(f.verbose_name).title()] = f
        by_header[f.name] = f
    # The PK column header from the export
    by_header[str(opts.pk.verbose_name).title()] = opts.pk
    by_header["Id"] = opts.pk
    by_header["id"] = opts.pk
    return by_header


def import_xlsx(model, file_obj):
    """
    Import the rows in ``file_obj`` into ``model``.

    Returns ``(created_count, updated_count, errors)``. ``errors`` is a list of
    ``"Row N: <message>"`` strings; if non-empty, the entire import is rolled
    back and ``(0, 0, errors)`` is returned.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as e:
        raise ImportError(f"Could not read Excel file: {e}") from e

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    headers = next(rows_iter, None)
    if not headers:
        raise ImportError("The spreadsheet is empty.")

    field_map = _build_field_map(model)
    opts = model._meta

    # Resolve each header to a field (or None to skip the column)
    columns = []
    pk_col_idx = None
    for i, header in enumerate(headers):
        if header is None:
            columns.append(None)
            continue
        field = field_map.get(str(header).strip())
        if field is None:
            columns.append(None)
            continue
        if field.name in _SKIP_FIELDS:
            columns.append(None)
            continue
        if isinstance(field, (FileField, ImageField)):
            # The export writes a URL; we can't reliably re-import that as a file.
            columns.append(None)
            continue
        columns.append(field)
        if field is opts.pk:
            pk_col_idx = i

    created_count = 0
    updated_count = 0
    errors = []

    with transaction.atomic():
        for row_num, row in enumerate(rows_iter, start=2):
            # Skip fully-empty rows
            if row is None or all(cell in (None, "") for cell in row):
                continue

            try:
                # ── Find existing row by PK ──────────────────────────────
                instance = None
                if pk_col_idx is not None and pk_col_idx < len(row):
                    pk_cell = row[pk_col_idx]
                    if pk_cell not in (None, ""):
                        try:
                            pk_val = int(pk_cell)
                            instance = model._default_manager.filter(
                                pk=pk_val).first()
                        except (ValueError, TypeError):
                            pass

                # ── Parse fields ─────────────────────────────────────────
                regular_data = {}
                m2m_data = {}
                for i, field in enumerate(columns):
                    if field is None or field is opts.pk:
                        continue
                    cell = row[i] if i < len(row) else None
                    if isinstance(field, ManyToManyField):
                        m2m_data[field.name] = _parse_m2m(field, cell)
                        continue
                    # Skip empty cells: leaves field unchanged on UPDATE,
                    # uses default on CREATE
                    if cell is None or str(cell).strip() == "":
                        continue
                    regular_data[field.name] = _parse_value(field, cell)

                # ── Create or update ─────────────────────────────────────
                if instance is None:
                    instance = model._default_manager.create(**regular_data)
                    created_count += 1
                else:
                    for k, v in regular_data.items():
                        setattr(instance, k, v)
                    instance.save()
                    updated_count += 1

                # M2Ms must be set after instance has a PK
                for field_name, related_objs in m2m_data.items():
                    getattr(instance, field_name).set(related_objs)

            except Exception as e:
                errors.append(f"Row {row_num}: {e}")

        if errors:
            transaction.set_rollback(True)
            return 0, 0, errors

    return created_count, updated_count, errors


# ── Admin view + monkey-patch ────────────────────────────────────────────────

class _UploadForm(forms.Form):
    xlsx_file = forms.FileField(
        label="Excel file (.xlsx)",
        widget=forms.FileInput(attrs={"accept": ".xlsx"}),
    )

    def clean_xlsx_file(self):
        f = self.cleaned_data["xlsx_file"]
        if not f.name.lower().endswith(".xlsx"):
            raise forms.ValidationError("Please upload a .xlsx file.")
        return f


def _import_view(modeladmin, request):
    """Admin view: GET shows the upload form, POST runs the import."""
    opts = modeladmin.model._meta
    changelist_url = reverse(
        f"admin:{opts.app_label}_{opts.model_name}_changelist"
    )

    if request.method == "POST":
        form = _UploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                created, updated, errors = import_xlsx(
                    modeladmin.model, form.cleaned_data["xlsx_file"]
                )
            except ImportError as e:
                messages.error(request, str(e))
                return HttpResponseRedirect(request.path)
            except Exception as e:  # noqa: BLE001 — last-resort safety net
                messages.error(request, f"Import failed: {e}")
                return HttpResponseRedirect(request.path)

            if errors:
                # Show up to 20 errors so the message bar isn't overwhelming
                for err in errors[:20]:
                    messages.error(request, err)
                if len(errors) > 20:
                    messages.error(
                        request,
                        f"… and {len(errors) - 20} more errors. "
                        "Nothing was saved (transaction rolled back).",
                    )
                else:
                    messages.error(
                        request,
                        "Import aborted — nothing was saved "
                        "(transaction rolled back).",
                    )
                return HttpResponseRedirect(request.path)

            messages.success(
                request,
                f"Import complete: {created} created, {updated} updated.",
            )
            return HttpResponseRedirect(changelist_url)
    else:
        form = _UploadForm()

    context = {
        **modeladmin.admin_site.each_context(request),
        "opts": opts,
        "form": form,
        "title": f"Import {opts.verbose_name_plural} from Excel",
        "changelist_url": changelist_url,
        "has_view_permission": True,
    }
    return render(request, "admin/import_xlsx.html", context)


# Monkey-patch ModelAdmin.get_urls so every registered admin gets the
# import-excel/ URL automatically. Done once at module import.
_orig_get_urls = admin.ModelAdmin.get_urls


def _patched_get_urls(self):
    urls = _orig_get_urls(self)
    info = self.model._meta.app_label, self.model._meta.model_name
    custom = [
        path(
            "import-excel/",
            self.admin_site.admin_view(
                lambda request, _self=self: _import_view(_self, request)
            ),
            name=f"{info[0]}_{info[1]}_import_xlsx",
        ),
    ]
    return custom + urls


admin.ModelAdmin.get_urls = _patched_get_urls
