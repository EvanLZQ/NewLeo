from __future__ import annotations  # 允许类型注解使用“前向引用”

import os  # 处理文件名/扩展名
import zipfile  # 读取图片 zip
from dataclasses import dataclass  # 返回结构化结果（导入统计）
from decimal import Decimal  # 金额字段更安全
from typing import Any, Dict, Iterable, List, Optional, Tuple  # 类型标注

from django.apps import apps  # 动态获取 Supplier 模型，避免循环 import
from django.core.exceptions import ValidationError  # 用于把“可读错误”抛回 Admin
from django.core.files.base import ContentFile  # 把 zip 内文件字节写入 Django storage
from django.core.files.storage import default_storage  # 使用 Django 默认存储（本地/云都兼容）
from django.db import transaction  # 保证数据库操作的原子性
from django.utils.text import slugify  # 生成 slug
from uuid import uuid4  # 生成唯一文件名，避免图片覆盖

from Product.models import ProductInfo, ProductInstance, ProductImage, ProductTag, ProductColorImg  # 导入本 app 模型

# =========================
#  返回给 Admin 的导入结果（便于 messages 展示）
# =========================


@dataclass(frozen=True)
class ImportResult:
    productinfo_created: int  # 新建 ProductInfo 数量
    productinfo_updated: int  # 更新 ProductInfo 数量
    instance_created: int  # 新建 ProductInstance 数量
    instance_updated: int  # 更新 ProductInstance 数量
    images_created: int  # 新建 ProductImage 数量
    missing_suppliers: List[str]  # Excel 里出现但 DB 找不到的 supplier 名字
    missing_tags: List[str]  # Excel 里出现但 DB 找不到的 tag 名字
    missing_images: List[str]  # Excel 指向但 zip 里找不到的图片路径（或文件名）


# =========================
#  读取 Excel：使用 openpyxl（Django 常用，稳定）
# =========================
def _read_excel_sheet_as_dicts(workbook, sheet_name: str) -> List[Dict[str, Any]]:
    """把指定 sheet 读成 list[dict]，每行 dict 的 key 是 header。"""  # docstring 让维护者快速理解

    if sheet_name not in workbook.sheetnames:  # 校验 sheet 是否存在
        raise ValidationError(
            f"Excel 缺少 sheet: {sheet_name}")  # 直接抛可读错误给 Admin

    ws = workbook[sheet_name]  # 取 worksheet
    rows = list(ws.iter_rows(values_only=True))  # 读取所有行（values_only=True 更快）

    if not rows:  # 空表
        return []  # 返回空列表

    # 第一行是列名
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    header = [h for h in header if h]  # 去掉空 header（防止尾部空列影响）

    data: List[Dict[str, Any]] = []  # 存放最终行数据

    for r in rows[1:]:  # 从第二行开始是真正数据
        row_values = list(r[: len(header)])  # 只取 header 对应长度，忽略多余空列
        row_dict = {header[i]: row_values[i]
                    for i in range(len(header))}  # 转 dict
        #  过滤掉“整行为空”的行（防止 Excel 尾部空行）
        if any(v is not None and str(v).strip() != "" for v in row_dict.values()):
            data.append(row_dict)  # 收集有效行

    return data  # 返回数据列表


def _normalize_bool(value: Any) -> bool:
    """把 Excel 可能出现的 True/False/1/0/'true'/'false' 统一转为 bool。"""  # 兼容性更强

    if isinstance(value, bool):  # 已经是 bool
        return value  # 直接返回
    if value is None:  # 空值按 False
        return False  # 默认 False
    s = str(value).strip().lower()  # 统一转字符串小写
    return s in {"1", "true", "yes", "y", "t"}  # 常见真值集合


def _normalize_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    """把 Excel 的数值转为 Decimal，空值则返回 default。"""  # 金额/价格字段用 Decimal 更稳

    if value is None or str(value).strip() == "":  # 空值
        return default  # 返回默认
    try:
        return Decimal(str(value))  # 通过字符串转 Decimal，避免 float 精度问题
    except Exception:
        raise ValidationError(f"无法把数值转换为 Decimal：{value}")  # 给出明确错误


def _normalize_int(value: Any, default: int = 0) -> int:
    """把 Excel 数值转 int，空值则返回 default。"""  # 尺寸/库存字段用

    if value is None or str(value).strip() == "":  # 空值
        return default  # 返回默认
    try:
        return int(value)  # 直接转 int
    except Exception:
        raise ValidationError(f"无法把数值转换为 int：{value}")  # 明确报错


def _ensure_unique_slug(base_slug: str) -> str:
    """保证 ProductInstance.slug 唯一；若冲突则追加 -2 / -3 ..."""  # 防止 unique 约束报错

    slug = base_slug  # 初始 slug
    i = 2  # 从 2 开始编号（更直观）
    while ProductInstance.objects.filter(slug=slug).exists():  # 若已存在则循环加后缀
        slug = f"{base_slug}-{i}"  # 追加后缀
        i += 1  # 继续递增
    return slug  # 返回唯一 slug


def _normalize_image_type(value: Any) -> str:
    """把 Excel 的 Carousel/Detail 等值规范为模型 choices: 'carousel'/'detail'。"""  # 与 ProductImage.image_type 对齐

    if value is None:  # 空值默认 carousel
        return "carousel"  # 默认轮播
    s = str(value).strip().lower()  # 小写
    if s in {"carousel", "car"}:  # 常见写法
        return "carousel"  # 返回
    if s in {"detail", "details", "det"}:  # 常见写法
        return "detail"  # 返回
    #  如果 Excel 写了不认识的值，宁可报错让 Admin 修正（避免脏数据）
    raise ValidationError(f"未知的 image_type：{value}（期望 Carousel/Detail）")


def _split_csv_cell(value: Any) -> List[str]:
    """把 Excel 单元格里逗号分隔的字符串拆成 list，并 strip。"""  # 用于 product_tag 和 images

    if value is None:  # 空值返回空列表
        return []  # 返回
    s = str(value).strip()  # 转字符串
    if not s:  # 空字符串
        return []  # 返回
    return [x.strip() for x in s.split(",") if x.strip()]  # 拆分并去空


def _build_zip_index(zip_file) -> Tuple[Dict[str, bytes], Dict[str, str]]:
    """
    读取 zip，返回：
    - path_to_bytes：zip 内完整路径 -> 文件 bytes
    - basename_to_path：文件名 -> zip 内路径（用于容错匹配）
    """  # 兼容 Excel 写“全路径”或“仅文件名”

    path_to_bytes: Dict[str, bytes] = {}  # 完整路径索引
    basename_to_path: Dict[str, str] = {}  # 文件名索引

    try:
        zf = zipfile.ZipFile(zip_file)  # 打开 zip
    except Exception:
        raise ValidationError("图片文件不是有效的 zip")  # 明确错误

    for name in zf.namelist():  # 遍历 zip 内文件
        if name.endswith("/"):  # 跳过目录条目
            continue  # 继续
        normalized = name.replace("\\", "/").lstrip("/")  # 统一路径分隔符
        data = zf.read(name)  # 读取文件 bytes
        path_to_bytes[normalized] = data  # 放入完整路径索引
        basename = os.path.basename(normalized)  # 取文件名
        #  若同名文件出现多次，这里保留第一个（也可以改成报错）
        basename_to_path.setdefault(basename, normalized)  # 建立文件名索引

    return path_to_bytes, basename_to_path  # 返回两种索引


def import_products_from_excel_and_zip(
    *,
    excel_file,  # request.FILES['excel_file']
    images_zip_file,  # request.FILES['images_zip_file']
    replace_existing_images: bool,  # 是否替换已有图片
) -> ImportResult:
    """
     主入口：解析 Excel，创建/更新 ProductInfo + ProductInstance，绑定 Supplier + ProductTag，
    并按 ProductImage sheet + zip 创建 ProductImage。
    """  # 强 docstring 方便未来维护

    #  延迟获取 Supplier 模型：避免 Product app 与 Supplier app 的 import 循环
    SupplierInfo = apps.get_model(
        "Supplier", "SupplierInfo")  # 'Supplier.SupplierInfo'

    #  用 openpyxl 读取 Excel（pandas 也可以，但 openpyxl 更“Django/后端风格”）
    from openpyxl import load_workbook  # 放函数内：避免未安装时报错影响其他模块加载

    try:
        # data_only=True 读取公式结果
        workbook = load_workbook(excel_file, data_only=True)
    except Exception:
        raise ValidationError("Excel 文件无法读取（请确认是 .xlsx）")  # 明确错误

    product_rows = _read_excel_sheet_as_dicts(
        workbook, "Product")  # 读取 Product sheet
    image_rows = _read_excel_sheet_as_dicts(
        workbook, "ProductImage")  # 读取 ProductImage sheet

    #  基础校验：至少要有数据
    if not product_rows:  # 没有产品行
        raise ValidationError("Excel 的 Product sheet 没有数据")  # 抛错提示

    #  读取 zip（如果不需要导入图片，也可以允许为空；但你需求写了要上传 zip）
    zip_path_to_bytes, zip_basename_to_path = _build_zip_index(
        images_zip_file)  # 建立 zip 索引

    #  把 ProductImage sheet 预处理成：sku -> list[图片条目]
    images_by_sku: Dict[str, List[Dict[str, Any]]] = {}  # sku 映射

    for r in image_rows:  # 遍历图片行
        sku = str(r.get("sku", "")).strip()  # sku 必填
        if not sku:  # 跳过无 sku 的行
            continue  # 继续
        images_by_sku.setdefault(sku, []).append(r)  # 追加到 sku 对应列表

    #  导入统计与“缺失清单”（用于 Admin message）
    productinfo_created = 0  # 统计新建 ProductInfo
    productinfo_updated = 0  # 统计更新 ProductInfo
    instance_created = 0  # 统计新建 ProductInstance
    instance_updated = 0  # 统计更新 ProductInstance
    images_created = 0  # 统计新建 ProductImage
    missing_suppliers: List[str] = []  # 缺失 supplier 名单
    missing_tags: List[str] = []  # 缺失 tag 名单
    missing_images: List[str] = []  # 缺失图片名单

    #  缓存（减少 DB 查询）
    supplier_cache: Dict[str, Any] = {}  # supplier_name -> SupplierInfo
    # tag_name(lower) -> ProductTag or None
    tag_cache: Dict[str, Optional[ProductTag]] = {}
    # color_display_name(lower) -> ProductColorImg or None
    colorimg_cache: Dict[str, Optional[ProductColorImg]] = {}

    #  先做一个“sku -> ProductInstance”映射（后续导图片用）
    instance_by_sku: Dict[str, ProductInstance] = {}  # sku -> instance

    #  数据库部分：用事务保证一致性（DB 成功后再导图片）
    with transaction.atomic():  # 原子事务开始

        for row in product_rows:  # 遍历 Product sheet 每一行
            # -------------------------
            #  1) 解析/校验关键字段
            # -------------------------
            # supplier 名字（你说 DB 里已存在）
            supplier_name = str(row.get("supplier", "")).strip()
            model_number = str(row.get("model_number", "")
                               ).strip()  # model_number
            sku = str(row.get("sku", "")).strip()  # sku（ProductInstance 唯一）

            if not supplier_name:  # supplier 必填
                raise ValidationError("Product sheet 有行缺少 supplier")  # 立刻报错
            if not model_number:  # model_number 必填
                raise ValidationError(
                    "Product sheet 有行缺少 model_number")  # 立刻报错
            if not sku:  # sku 必填
                raise ValidationError("Product sheet 有行缺少 sku")  # 立刻报错

            # -------------------------
            #  2) 获取 SupplierInfo（按名字查）
            # -------------------------
            if supplier_name not in supplier_cache:  # 缓存未命中
                #  这里假设 SupplierInfo 有字段 name；如果你实际字段叫 title/company_name，请改这里即可
                supplier_obj = SupplierInfo.objects.filter(
                    name__iexact=supplier_name).first()  # 不区分大小写
                supplier_cache[supplier_name] = supplier_obj  # 写入缓存

            supplier_obj = supplier_cache[supplier_name]  # 从缓存取

            if supplier_obj is None:  # DB 找不到 supplier
                if supplier_name not in missing_suppliers:  # 避免重复记录
                    missing_suppliers.append(supplier_name)  # 加入缺失列表
                #  supplier 缺失属于硬错误（否则 ProductInfo 外键无法正确设置）
                raise ValidationError(
                    f"数据库找不到 Supplier：{supplier_name}")  # 中断导入

            # -------------------------
            #  3) Upsert ProductInfo（以 supplier + model_number 作为“自然键”）
            # -------------------------
            productinfo_defaults = {  # 用 Excel 的字段填充 ProductInfo
                # 名称（允许为空）
                "name": (row.get("name") or "").strip() if isinstance(row.get("name"), str) else (row.get("name") or ""),
                # rmb_price
                "rmb_price": _normalize_decimal(row.get("rmb_price"), Decimal("0")),
                # price
                "price": _normalize_decimal(row.get("price"), Decimal("0")),
                # description（TextField）
                "description": row.get("description") or "",
                # letter_size（按你模型 choices）
                "letter_size": row.get("letter_size") or "M",
                "string_size": row.get("string_size") or "",  # string_size
                # frame_weight
                "frame_weight": _normalize_int(row.get("frame_weight"), 0),
                "bifocal": _normalize_bool(row.get("bifocal")),  # bifocal
                # frame_width
                "frame_width": _normalize_int(row.get("frame_width"), 0),
                # lens_width
                "lens_width": _normalize_int(row.get("lens_width"), 0),
                "bridge": _normalize_int(row.get("bridge"), 0),  # bridge
                # temple_length
                "temple_length": _normalize_int(row.get("temple_length"), 0),
                # lens_height
                "lens_height": _normalize_int(row.get("lens_height"), 0),
                # upper_wearable_width
                "upper_wearable_width": _normalize_int(row.get("upper_wearable_width"), 0),
                # lower_wearable_width
                "lower_wearable_width": _normalize_int(row.get("lower_wearable_width"), 0),
                "gender": row.get("gender") or "Unisex",  # gender
                "nose_pad": row.get("nose_pad") or "Standard",  # nose_pad
                # frame_style
                "frame_style": row.get("frame_style") or "Full-Rim",
                # pd_upper_range
                "pd_upper_range": _normalize_int(row.get("pd_upper_range"), 80),
                # pd_lower_range
                "pd_lower_range": _normalize_int(row.get("pd_lower_range"), 30),
            }  # defaults 字典结束

            product_info, created = ProductInfo.objects.update_or_create(  # Upsert：存在则更新，不存在则创建
                supplier=supplier_obj,  # 外键 supplier
                model_number=model_number,  # model_number
                defaults=productinfo_defaults,  # 其他字段
            )  # update_or_create 结束

            if created:  # 新建计数
                productinfo_created += 1  # +1
            else:  # 更新计数
                productinfo_updated += 1  # +1

            # -------------------------
            #  4) 绑定 ProductTag（Excel 的 product_tag：逗号分隔）
            # -------------------------
            tag_names = _split_csv_cell(
                row.get("product_tag"))  # 拆分成 list[str]

            for tag_name in tag_names:  # 遍历每个 tag
                key = tag_name.strip().lower()  # 统一 key
                if key not in tag_cache:  # 缓存未命中
                    tag_cache[key] = ProductTag.objects.filter(
                        name__iexact=tag_name.strip()).first()  # 按名字查
                tag_obj = tag_cache[key]  # 从缓存取
                if tag_obj is None:  # DB 找不到 tag
                    if tag_name not in missing_tags:  # 避免重复记录
                        missing_tags.append(tag_name)  # 记录缺失
                    continue  # 继续导入（tag 缺失不阻断：避免整个导入失败）
                #  ProductTag 模型里是：product = ManyToManyField(ProductInfo)
                tag_obj.product.add(product_info)  # 绑定关系（幂等）

            # -------------------------
            #  5) Upsert ProductInstance（以 sku 作为自然键）
            # -------------------------
            #  slug：如果已有 instance 保留原 slug；如果新建或 slug 为空，则生成
            # slugify 可能返回空（比如全中文），兜底用 sku
            proposed_slug = slugify(sku) or sku.lower()
            unique_slug = _ensure_unique_slug(proposed_slug)  # 确保唯一

            #  尝试按 color_display_name 找到 ProductColorImg（如果你 DB 里有）
            color_display_name = str(
                row.get("color_display_name", "")).strip()  # 颜色展示名
            color_key = color_display_name.lower()  # 统一 key
            if color_key and color_key not in colorimg_cache:  # 缓存未命中
                colorimg_cache[color_key] = ProductColorImg.objects.filter(
                    title__iexact=color_display_name).first()  # 按 title 查
            color_img_obj = colorimg_cache.get(
                color_key) if color_key else None  # 取出 color_img（可为空）

            instance_defaults = {  # instance 字段 defaults
                "product": product_info,  # 外键：关联 ProductInfo
                "slug": unique_slug,  # slug（unique）
                "stock": _normalize_int(row.get("stock"), 0),  # stock
                # reduced_price
                "reduced_price": _normalize_decimal(row.get("reduced_price"), Decimal("0")),
                # instance.price（允许你以后区分）
                "price": _normalize_decimal(row.get("price"), Decimal("0")),
                "color_img": color_img_obj,  # color_img（可为空）
                # color_base_name（模型 choices 里是小写）
                "color_base_name": str(row.get("color_base_name") or "black").strip().lower(),
                "color_display_name": color_display_name or "Default",  # color_display_name
                # instance 描述来自 description_ins
                "description": row.get("description_ins") or "",
                "online": _normalize_bool(row.get("online")),  # online
            }  # defaults 结束

            instance_obj, instance_created_flag = ProductInstance.objects.update_or_create(  # Upsert instance
                sku=sku,  # sku 唯一
                defaults=instance_defaults,  # 更新字段
            )  # update_or_create 结束

            #  如果 instance 已存在并且 slug 不为空，通常不建议每次导入都改 slug（会影响 URL）
            if not instance_created_flag and instance_obj.slug and instance_obj.slug != unique_slug:  # slug 已存在且不同
                #  这里选择“保留旧 slug”，所以把 defaults 的 slug 覆盖回旧值（避免 URL 变化）
                #  你若希望 slug 永远跟随 sku，可删除这段逻辑
                pass  # 明确不做任何事（保留旧 slug）

            if instance_created_flag:  # 新建计数
                instance_created += 1  # +1
            else:  # 更新计数
                instance_updated += 1  # +1

            instance_by_sku[sku] = instance_obj  # 保存映射（后续导图片使用）

        #  如果用户选择替换图片：先删 DB 记录（文件是否删除取决于 storage 策略，通常建议异步清理）
        if replace_existing_images:  # 需要替换
            ProductImage.objects.filter(
                productInstance__in=instance_by_sku.values()).delete()  # 删除旧图片记录

        #  DB 成功提交后再写 storage & 创建 ProductImage（避免 DB rollback 产生“孤儿文件”）
        def _on_commit_import_images():  # on_commit 回调函数
            nonlocal images_created  # 修改外层统计变量
            nonlocal missing_images  # 修改外层缺失列表

            # 收集要 bulk_create 的 ProductImage
            new_image_objs: List[ProductImage] = []

            for sku, instance_obj in instance_by_sku.items():  # 遍历每个实例
                for img_row in images_by_sku.get(sku, []):  # 取该 sku 的图片行
                    image_paths = _split_csv_cell(
                        img_row.get("images"))  # 逗号分隔图片路径
                    alt = (img_row.get("alt") or "").strip() if isinstance(
                        img_row.get("alt"), str) else (img_row.get("alt") or "")  # alt 文本
                    image_type = _normalize_image_type(
                        img_row.get("image_type"))  # 规范化 image_type

                    for p in image_paths:  # 遍历每个图片 path
                        normalized = p.replace("\\", "/").lstrip("/")  # 路径规范化

                        data = None  # 图片 bytes（默认空）
                        filename = os.path.basename(normalized)  # 默认用 basename

                        #  先尝试“完整路径匹配”
                        if normalized in zip_path_to_bytes:  # 完整路径命中
                            data = zip_path_to_bytes[normalized]  # 取 bytes
                        else:
                            #  再尝试“仅文件名匹配”（容错：Excel 可能没写目录）
                            fallback_path = zip_basename_to_path.get(
                                filename)  # 用 basename 查
                            if fallback_path:  # 找到
                                data = zip_path_to_bytes.get(
                                    fallback_path)  # 取 bytes

                        if data is None:  # zip 里找不到图片
                            missing_images.append(normalized)  # 记录缺失
                            continue  # 跳过

                        #  生成不冲突的 storage 文件名（避免覆盖）
                        _, ext = os.path.splitext(filename)  # 提取扩展名
                        ext = ext or ".jpg"  # 没扩展名就兜底
                        # 统一存到 import 子目录
                        storage_name = f"product_images/import/{uuid4().hex}{ext}"
                        saved_path = default_storage.save(
                            storage_name, ContentFile(data))  # 写入 storage，返回最终路径

                        img_obj = ProductImage(  # 构建 ProductImage（先不 save）
                            productInstance=instance_obj,  # 关联 instance
                            alt=alt or instance_obj.sku,  # alt：为空就用 sku
                            image_type=image_type,  # image_type
                        )  # 构建结束
                        img_obj.image.name = saved_path  # 直接设置 image 字段的存储路径（避免逐个 save）
                        new_image_objs.append(img_obj)  # 收集起来 bulk_create

            if new_image_objs:  # 有新图片才写 DB
                ProductImage.objects.bulk_create(
                    new_image_objs, batch_size=500)  # 批量插入更快
                images_created += len(new_image_objs)  # 更新统计

        transaction.on_commit(_on_commit_import_images)  # 注册“事务提交后执行”的回调

    #  事务结束：如果成功提交，on_commit 会执行并写入图片与 ProductImage

    return ImportResult(  # 返回统计结果给 Admin
        productinfo_created=productinfo_created,  # 回填
        productinfo_updated=productinfo_updated,  # 回填
        instance_created=instance_created,  # 回填
        instance_updated=instance_updated,  # 回填
        images_created=images_created,  # 回填
        missing_suppliers=missing_suppliers,  # 回填
        missing_tags=missing_tags,  # 回填
        missing_images=missing_images,  # 回填
    )  # 返回结束
