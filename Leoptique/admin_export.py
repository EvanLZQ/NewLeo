"""
Generic site-wide "Export to Excel (.xlsx)" admin action.

Importing this module auto-registers the action via
``admin.site.add_action(...)`` so it appears on every ModelAdmin's action
dropdown without needing per-model code changes.

The action serialises the queryset Django passes (selected rows, or all
filtered rows if the user clicked "Select all X items") to an .xlsx file
and returns it as a download.
"""
import json
from datetime import date, datetime, time
from decimal import Decimal

from django.contrib import admin
from django.db.models import (
    DateField,
    DateTimeField,
    FileField,
    ForeignKey,
    ImageField,
    JSONField,
    ManyToManyField,
    OneToOneField,
    TimeField,
)
from django.http import HttpResponse


# ── Field-level serialization ────────────────────────────────────────────────

def _serialize_value(field, instance):
    """
    Convert a single field's value on ``instance`` into an Excel-friendly
    string/number. Returns "" for None.
    """
    # ManyToMany: comma-separated str() of related objects
    if isinstance(field, ManyToManyField):
        related = getattr(instance, field.name).all()
        return ", ".join(str(o) for o in related)

    # Choices field: prefer human-readable label
    if getattr(field, "choices", None):
        getter = getattr(instance, f"get_{field.name}_display", None)
        if callable(getter):
            value = getter()
            return "" if value is None else str(value)

    # Foreign keys / O2O: use str(related_obj)
    if isinstance(field, (ForeignKey, OneToOneField)):
        related = getattr(instance, field.name, None)
        return "" if related is None else str(related)

    # All other fields: read the raw attribute
    value = getattr(instance, field.name, None)
    if value is None or value == "":
        return ""

    if isinstance(field, (FileField, ImageField)):
        try:
            return value.url
        except (ValueError, AttributeError):
            # File field with no file attached
            try:
                return value.name or ""
            except AttributeError:
                return ""

    if isinstance(field, JSONField):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except (TypeError, ValueError):
            return str(value)

    if isinstance(field, DateTimeField) and isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(field, DateField) and isinstance(value, date):
        return value.isoformat()
    if isinstance(field, TimeField) and isinstance(value, time):
        return value.isoformat()

    if isinstance(value, Decimal):
        return str(value)

    return str(value)


# ── Action callback ──────────────────────────────────────────────────────────

# Sensitive field names skipped in the export, regardless of model.
_SENSITIVE_FIELDS = {"password"}


@admin.action(description="Export selected to Excel (.xlsx)")
def export_as_xlsx(modeladmin, request, queryset):
    """Generic admin action: dump the queryset to an .xlsx download."""
    # Local import so test suites / management commands that don't need
    # openpyxl don't pay the import cost on Django startup.
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    model = queryset.model
    opts = model._meta

    # Build column list: concrete fields + M2Ms, skipping sensitive fields
    fields = [
        f for f in list(opts.fields) + list(opts.many_to_many)
        if f.name not in _SENSITIVE_FIELDS
    ]

    # Pre-fetch FKs and M2Ms to avoid N+1 queries during serialization
    fk_names = [
        f.name for f in opts.fields
        if isinstance(f, (ForeignKey, OneToOneField))
        and f.name not in _SENSITIVE_FIELDS
    ]
    m2m_names = [f.name for f in opts.many_to_many]
    if fk_names:
        queryset = queryset.select_related(*fk_names)
    if m2m_names:
        queryset = queryset.prefetch_related(*m2m_names)

    # Build the workbook
    wb = Workbook()
    ws = wb.active
    # Excel limits sheet titles to 31 chars
    ws.title = (str(opts.verbose_name_plural) or opts.model_name)[:31]

    # Header row
    headers = [str(f.verbose_name).title() for f in fields]
    ws.append(headers)

    # Data rows
    for instance in queryset.iterator():
        row = [_serialize_value(f, instance) for f in fields]
        ws.append(row)

    # Basic column widths for readability
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(
            12, min(50, len(header) + 2)
        )

    # Build the HTTP response
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"{opts.app_label}_{opts.model_name}_{timestamp}.xlsx"
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── Site-wide registration (runs once on import) ─────────────────────────────

admin.site.add_action(export_as_xlsx, "export_as_xlsx")
